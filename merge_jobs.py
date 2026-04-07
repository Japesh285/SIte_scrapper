#!/usr/bin/env python3
"""Merge job detail JSON files into a single Excel spreadsheet.

Extracts the per-job schema from every scrape_result_*.json
and writes one flat Excel with one row per job.

Usage:
    python3 merge_jobs.py
"""

import json
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent / "job-details"

HEADERS = [
    "id", "title", "company_name", "job_link",
    "experience", "locations", "educational_qualifications",
    "required_skill_set", "remote_type", "posted_on",
    "salary", "is_active", "first_seen", "last_seen",
    "additional_sections",
    "scrap_url", "scrap_strategy", "scrap_parser",
    "scrap_confidence", "scrap_ai_forced",
    "ai_input_tokens", "ai_output_tokens", "ai_total_tokens",
]


def main():
    files = sorted(BASE.rglob("scrape_result_*.json"))
    if not files:
        print("No job files found under job-details/")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for f in files:
        with open(f, "r", encoding="utf-8") as fh:
            data = json.load(fh)

        for job in data.get("jobs", []):
            scrap = job.get("Scrap_json", {})
            ai = job.get("ai_usage", {})

            ws.cell(row=row, column=1, value=job.get("id", ""))
            ws.cell(row=row, column=2, value=job.get("title", ""))
            ws.cell(row=row, column=3, value=job.get("company_name", ""))
            ws.cell(row=row, column=4, value=job.get("job_link", ""))
            ws.cell(row=row, column=5, value=job.get("experience", ""))
            ws.cell(row=row, column=6, value=json.dumps(job.get("locations", [])))
            ws.cell(row=row, column=7, value=str(job.get("educational_qualifications", "")))
            ws.cell(row=row, column=8, value=", ".join(job.get("required_skill_set", [])))
            ws.cell(row=row, column=9, value=job.get("remote_type", ""))
            ws.cell(row=row, column=10, value=job.get("posted_on", ""))
            ws.cell(row=row, column=11, value=job.get("salary", ""))
            ws.cell(row=row, column=12, value=job.get("is_active", ""))
            ws.cell(row=row, column=13, value=job.get("first_seen", ""))
            ws.cell(row=row, column=14, value=job.get("last_seen", ""))
            ws.cell(row=row, column=15, value=json.dumps(job.get("additional_sections", [])))
            ws.cell(row=row, column=16, value=scrap.get("url", ""))
            ws.cell(row=row, column=17, value=scrap.get("strategy", ""))
            ws.cell(row=row, column=18, value=scrap.get("parser_used", ""))
            ws.cell(row=row, column=19, value=scrap.get("confidence", ""))
            ws.cell(row=row, column=20, value=scrap.get("ai_forced", ""))
            ws.cell(row=row, column=21, value=ai.get("input_tokens", 0))
            ws.cell(row=row, column=22, value=ai.get("output_tokens", 0))
            ws.cell(row=row, column=23, value=ai.get("total_tokens", 0))
            row += 1

    output = Path(__file__).parent / "merged_jobs.xlsx"
    wb.save(output)
    print(f"Saved {row - 2} jobs from {len(files)} files to {output}")


if __name__ == "__main__":
    main()
